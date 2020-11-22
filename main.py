import sys
import openpyxl
import os
import PyPDF2
from pathlib import Path

source_list = []


def main():
    source_path = get_source_link()
    target_path = get_target_link()
    print(show_source_to_user(source_path))
    print(page_numb_target(target_path))


def get_source_link():
    #path = input("Insert path to source: ")
    source_path = "/home/jkafka/qa_finder_files/source.xlsx"
    return source_path


def get_target_link():
    #path = input("Insert path to target: ")
    target_path = "/home/jkafka/qa_finder_files/PN12345_A_view.pdf"
    return target_path

def source_max_row_get(source_path):
    wb_obj = openpyxl.load_workbook(source_path)
    sheet_obj = wb_obj.active
    max_row = sheet_obj.max_row
    return max_row, sheet_obj


def source_list_data_get(source_path):
    max_row, sheet_obj = source_max_row_get(source_path)
    for i in range(1, max_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=2)
        source_list.append(cell_obj.value)
    return source_list


def show_source_to_user(source_path):
    source_data = source_list_data_get(source_path)
    return source_data


def approve_source():
    pass


def get_target_from_pdf():
    pass


def get_pdf_name(target_path):
    base=os.path.basename(target_path)
    pdf_name = os.path.splitext(base)[0]
    return pdf_name


def page_numb_target(target_path):
    pdfFileObj = open(target_path, 'rb') 
    pdfReader = PyPDF2.PdfFileReader(pdfFileObj) 
    num_pages = pdfReader.numPages 
    return num_pages


def make_clean_text():
    pass


def source_and_target_compare():
    pass


def compare_result():
    pass


if __name__ == '__main__':
    main()
